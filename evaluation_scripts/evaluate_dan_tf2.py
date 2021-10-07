import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3' 
import tensorflow as tf
import cv2
import glob
import shutil
import pandas as pd
from tensorflow.keras.applications.xception import Xception
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import ImageDataGenerator

from sklearn.metrics import confusion_matrix, classification_report
from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



####################
# Helper Functions #
####################
def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:
    
    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws

# append df (data) to existing excel sheet: https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas#38075046
def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:

    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt
    filename = Path(filename)
    file_exists = filename.is_file()
    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        len(df.columns[col_no]) + 6)
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()


# finds h5 file with lowest validation loss
def find_lowest_val_loss(weights_dir):
    lowest_val_loss = 1000
    weight_name_best = ''
    for weightName in os.listdir(weights_dir):
        val_loss = weightName.split('-')[-1].split('.h5')[0]
        val_loss = float(val_loss)
        if(val_loss < lowest_val_loss):
            lowest_val_loss = val_loss
            weight_name_best = weightName
    return weight_name_best

#finds h5 file wit highest validation accuracy
def find_highest_val_acc(weights_dir):
    highest_val_acc = 0.5
    weight_name_best = ''
    for weightName in os.listdir(weights_dir):
        val_acc = weightName.split('-')[1]
        val_acc = float(val_acc)
        if(val_acc > highest_val_acc):
            highest_val_acc = val_acc
            weight_name_best = weightName
    return weight_name_best

#locates h5 file in the h5_to_uff folder
def find_h5_file(folder):
    for filepath in glob.glob(folder + '\*'):
        filename = filepath.split('\\')[-1]
        if '.h5' in filename:
            return filename

#returns all classification report information
def get_classification_report_info(CLASSES, report):
    train_images = []
    precision = []
    recall = []
    f1_score = []
    support = []
    for clss in CLASSES:
        train_images.append(train_count.count(CLASSES.index(clss)))
        precision.append(round(float(report[clss]['precision']),2))
        recall.append(round(float(report[clss]['recall']),2))
        f1_score.append(round(float(report[clss]['f1-score']),2))
        support.append(float(report[clss]['support']))
    accuracy = round(report['accuracy'],2)
    macro_avg_precision = round(report['macro avg']['precision'],2)
    macro_avg_recall = round(report['macro avg']['recall'],2)
    macro_avg_f1_score = round(report['macro avg']['f1-score'],2)
    macro_avg_support = round(report['macro avg']['support'],2)
    return [str(train_images), str(precision), str(recall), str(f1_score), str(support), accuracy, macro_avg_precision, macro_avg_recall, macro_avg_f1_score, macro_avg_support]



######################
# Specify GPU Device #
######################
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = '1'


##############
# Load Model #
##############
model_version = 49

weights_from_h5_folder = True
print(f"Loading Model v{model_version}")
if weights_from_h5_folder:
    weights_version = find_h5_file(r'D:\Daniel\PMD\scripts\H5 to UFF\pmd_v{}'.format(model_version))
    model_path = r"D:\Daniel\PMD\scripts\H5 to UFF\pmd_v{}\{}".format(model_version, weights_version)

else:
    weights_dir = r"D:\Daniel\PMD\model_training\weights_dan\PMD_V{}".format(model_version)
    weights_version = find_highest_val_acc(weights_dir)
    weights_version = find_lowest_val_loss(weights_dir)
    model_path = r"D:\Daniel\PMD\model_training\weights_dan\PMD_V{}\{}".format(model_version, weights_version)

print(weights_version)
model = load_model(model_path)


###############
# Directories #
###############
test_dir = r"D:\Daniel\PMD\model_training\final_data\test"
train_dir = r"D:\Daniel\PMD\model_training\final_data\combined_data_29.9.2021\train"
print_dir = r"D:\Daniel\PMD\model_training\wrong_classifications\PMD_V{}".format(model_version) #directory to print prediction of test images
excel_path = r"D:\Daniel\PMD\model_training\model_evaluation_dan.xlsx"



####################
# Data Preparation #
####################
CLASSES = ["cyclist", "ebike", 'motorcycle', 'pedestrian', 'standing_scooter']
print(CLASSES)

IMG_SIZE = (260,110)
train_count = []
test_data = []
y_true = []

for class_name in (CLASSES):
    print(class_name)
    for img_path in os.listdir(os.path.join(test_dir, class_name)):
        try:
            y_true.append(CLASSES.index(class_name))
            img = cv2.imread(os.path.join(test_dir,class_name, img_path))
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            img = cv2.resize(img, (IMG_SIZE[1],IMG_SIZE[0]), cv2.INTER_LANCZOS4)
            img = img/255.
            
            test_data.append(img)
        except:
            pass
    
    # count total number of train images
    for img_path in os.listdir(os.path.join(train_dir, class_name)):
        train_count.append(CLASSES.index(class_name))


# convert images into arrays
arr_ = np.array(test_data)
y_true =  np.array(y_true)

# reshape arrays based on image shape trained in model
arr = arr_.reshape(-1, IMG_SIZE[0],IMG_SIZE[1],3)



##############
# Evaluation #
##############
# predict test data
y_pred = model.predict(arr, verbose=1)
y_pred = list(np.argmax(y_pred,axis=1))

# Classification Report
print('Classification Report')
report = classification_report(y_true, y_pred, target_names = CLASSES, output_dict=True) #report to obtain information from
report_string = classification_report(y_true, y_pred, target_names = CLASSES) #report as a string to print out
print(report_string)

# Confusion Matrix
print('Confusion Matrix')
matrix = confusion_matrix(y_true, y_pred)
print(matrix)



################################
# Store Evaluation Information #
################################
# store classification report information inside excel sheet
store_report = True
if store_report:
    model_architecture = 'DenseNet201'
    notes = input('Note down approach to training: ')
    weights_name = weights_version
    train_images, precision, recall, f1_score, support, accuracy, macro_avg_precision, macro_avg_recall, macro_avg_f1_score, macro_avg_support = get_classification_report_info(CLASSES, report)

    df = pd.read_excel(excel_path)
    prev_model_name = df.tail(1)['Model Version'][df.shape[0]-1]
    current_model_name = 'PMD_V' + str(int(prev_model_name.split('PMD_V')[1]) + 1)

    report_info = [current_model_name, 
                    model_architecture,
                    train_images,
                    str(IMG_SIZE), 
                    precision, 
                    recall, 
                    f1_score, 
                    support, 
                    str([macro_avg_precision, macro_avg_recall, macro_avg_f1_score, macro_avg_support]), 
                    accuracy, 
                    [str(matrix[i]) for i in range(len(matrix))], 
                    weights_name]

    col_names = [] #get list of all column names
    for col_name, _ in df.iteritems():
        col_names.append(col_name) 

    df_dict = {} #get dictionary of classification report information
    for i in range(len(report_info)):
        if i == 10:
            matrix_rows = report_info[i]
            for j in range(len(CLASSES)):
                df_dict[col_names[i+j]] = matrix_rows[j]
        elif i == len(report_info)-1:
            df_dict['Remarks'] = report_info[i]
            df_dict['Notes'] = notes
        else:
            df_dict[col_names[i]]= report_info[i]

    df = pd.DataFrame(df_dict,index=[0])
    append_df_to_excel(excel_path, df, header=None, index=False) #appends dataframe to exisiting excel sheet
    excel_name = excel_path.split('\\')[-1]
    print(f'Stored evaluation information into {excel_name}')

    '''
    # this doesnt work for latest version of pandas
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') #ensure excel file is not open
    book = load_workbook(excel_path)
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # write out the new sheet
    df.to_excel(writer, index = False, header= False, startrow=writer.sheets['Sheet1'].max_row)
    writer.save()
    '''


# Print prediction of test data
print_test = False
if print_test:
    if not os.path.exists(print_dir):
        os.mkdir(print_dir)
    else:
        shutil.rmtree(print_dir)
        os.mkdir(print_dir)

    for i,(p,y) in enumerate(zip(y_pred, y_true)):
        curr_img = cv2.cvtColor(((arr_[i])*255).astype('uint8'), cv2.COLOR_BGR2RGB)
        folder_name = str(y)+str(p)
        folder_path = os.path.join(r"D:\Daniel\PMD\model_training\wrong_classifications\PMD_V{}".format(model_version), folder_name)
        if(not os.path.exists(folder_path)):
            os.mkdir(folder_path)
        img_name = str(len(os.listdir(folder_path)))+'.jpg'
        cv2.imwrite(os.path.join(folder_path, img_name), curr_img)
    print('Saved wrong classifications')
